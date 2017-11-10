import openpyxl
import docx
import sys
import os
'''
PI ID & Title: Col 4,
Activity ID: Col 5,
PI ID: Col 6,
TK ID: Col 7,
Work Type: Col 8,
Crop: Col 13,
Planned SD: Col 16,
Allocated SD: Col 17,
Planned $: Col 20,
Allocated $: Col 21,
Named Resource: Col 22,
Activity Comments: Col 23,
Portfolio: Col 24
'''

sourceFile = sys.argv[1]

#named resource report
nameSource = openpyxl.load_workbook(sourceFile)
sheets = nameSource.get_sheet_names()
source = sheets[0]
page = nameSource.get_sheet_by_name(source)
rowMax = page.max_row
staff = ['Beauchamps, Charles', 'Briois, Sebastien', 'Burthey, Kirkland', 'Cohn, Jonathan', 'Conte, Matthieu', 'El Mrani, Aicha',
         'Ganko, Eric', 'Graham, Steve', 'Green, Julie', 'Kavanaugh, Laura', 'Leysens, Maurice',  'Moughamer, Todd', 'Muller, Cedric',
         'NGOU, KISSI', 'Parvathaneni, Usha', 'Rose, Mark', 'Sharma, Pranav', 'Tiberi, Guillaume', 'Whalen, Brian']
ext = ['NCGR', 'Nebion']

#new blank report
kirkwb = openpyxl.Workbook()
newsheet = kirkwb.get_sheet_names()
kirksource = newsheet[0]
kirkpage = kirkwb.get_sheet_by_name(kirksource)

#Headers
staffTotal = []
count = 2
kirkpage['A1'] = str(page.cell(row=1, column=5).value)
kirkpage['B1'] = str(page.cell(row=1, column=6).value)
kirkpage['C1'] = str(page.cell(row=1, column=7).value)
kirkpage['D1'] = str(page.cell(row=1, column=8).value)
kirkpage['E1'] = str(page.cell(row=1, column=17).value)
kirkpage['F1'] = str(page.cell(row=1, column=21).value)
kirkpage['G1'] = str(page.cell(row=1, column=23).value)

#Employee loop
for j in range(0, len(staff)):
    kirkrowMax = kirkpage.max_row
    count = kirkrowMax + 1
    employeeTotal = []
    for i in range(1,rowMax):
        if str(page.cell(row=i, column=22).value) == staff[j]:
            if str(page.cell(row=i, column=17).value) != '-':
                employeeTotal.append(page.cell(row=i, column=17).value)
                staffTotal.append(page.cell(row=i, column=17).value)
            count += 1
            kirkpage['A' + str(kirkrowMax + 1)] = staff[j]
            kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
            kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
            kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
            kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
            kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
            kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
            kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
    kirkpage['E' + str(kirkrowMax + 1)] = sum(employeeTotal)
staffgrandTotal = sum(staffTotal)
kirkrowMax = kirkpage.max_row
kirkpage['A' + str(kirkrowMax + 1)] = 'FTE Grand Total'
kirkpage['E' + str(kirkrowMax + 1)] = str(staffgrandTotal)

#external provider loop

providerTotal = []
for k in range(0, len(ext)):
    kirkrowMax = kirkpage.max_row
    count = kirkrowMax + 1
    extTotal = []
    for i in range(1,rowMax):
        if str(page.cell(row=i, column=22).value) == ext[k]:
            if str(page.cell(row=i, column=21).value) != '-':
                extTotal.append(page.cell(row=i, column=21).value)
                providerTotal.append(page.cell(row=i, column=21).value)
            count += 1
            kirkpage['A' + str(kirkrowMax + 1)] = ext[k]
            kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
            kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
            kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
            kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
            kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
            kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
            kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
    kirkpage['F' + str(kirkrowMax + 1)] = sum(extTotal)
providergrandTotal = sum(providerTotal)
kirkrowMax = kirkpage.max_row
kirkpage['A' + str(kirkrowMax + 1)] = 'External Provider Grand Total'
kirkpage['F' + str(kirkrowMax + 1)] = str(providergrandTotal)



kirkwb.save('v2excelreport.xlsx')  

#! python3
#This script creates a dictionary of named resource data and SD allocations
#writes the dictionary to a python script 
#then outputs those numbers into an excel spreadsheet and chart

from time import strftime
import openpyxl, pprint
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
    BarChart
)
print('Opening workbook...')

#load data from resource report
wb = openpyxl.load_workbook('named_resource_breakdown4.xlsx')
sheet = wb.active
resourceData = {}
allNames = []

#populate dictionary with names and SD estimates
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    name = sheet['V' + str(row)].value
    if str(sheet['V' + str(row)].value) not in allNames:
        allNames.append(str(sheet['V' + str(row)].value))
    Activity = sheet['E' + str(row)].value
    SD = sheet['Q' + str(row)].value
    resourceData.setdefault(name, {'Total': 0})
    if SD != '-':
        resourceData[name]['Total'] += float(SD)

#print results to a python file which can be imported as a module
print('Writing results...')
resultFile = open('FTE_' + strftime('%m-%d-%y') +'.py', 'w')
resultFile.write(pprint.pformat(resourceData))
resultFile.close()
print('Done.')

#open new workbook to create chart
chartwb = openpyxl.Workbook()
namesheet = chartwb.get_sheet_names()
tsource = namesheet[0]
tsheet = chartwb.get_sheet_by_name(tsource)
 
#append totals to a list to generate chart
ttotals = []
for i in allNames:
    ttotals.append(resourceData[i]['Total'])

#headers for table that will create chart
count = 1
tsheet['A1'] = 'Name'
tsheet['B1'] = 'Total SD'

#populate data for table
for j in range (0, len(allNames)):
    count +=1
    tsheet['A' + str(count)] = allNames[j]
    tsheet['B' + str(count)] = ttotals[j]

#Generate and output chart
data = openpyxl.chart.Reference(tsheet, min_col=2, min_row=1, max_row=(len(allNames)+1))
titles = openpyxl.chart.Reference(tsheet, min_col=1, min_row=2, max_row=(len(allNames) +1))
#chart = BarChart3D()
chart = BarChart()
chart.style = 11
chart.type = 'bar'
chart.title = "Named Resource Total SDs"
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)
tsheet.add_chart(chart, 'E5')
chartwb.save('readsourcing.xlsx')

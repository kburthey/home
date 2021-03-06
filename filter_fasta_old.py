#!Python3
'''
fasta file parser
diplay name and read count for each item
sort output by read count total
'''


import sys, argparse, openpyxl, os
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import(
    Reference,
    Series,
    BarChart3D,
    BarChart
)
from Bio import SeqIO
from pprint import pprint

infile = sys.argv[1]

#New blank metadata
meta = openpyxl.Workbook()
meta_page = meta.create_sheet("Metadata",0)

#Headers
meta_page['A1'] = "infile"
meta_page['B1'] = "max"
meta_page['C1'] = "min"
meta_page['D1'] = "outfile"

#Global variables
low_count = {}
low_reads = []
normal_count = {}
normal_reads = []
high_count = {}
high_reads = []
results = {}
steps = []
groups = []
highest = 0
try:
    low_limit = int(sys.argv[2])
    upper_limit = int(sys.argv[3])
except ValueError:
    print("lower and upper limits must be integers")
    sys.exit(0)

#loop through each read and compare read length to upper and lower limits
records = list(SeqIO.parse(infile, "fasta"))
with open(infile + "_filtered_" + str(low_limit) + "_" + str(upper_limit) + ".fa", 'w') as target:
    for i in range(0, len(records)):
        if len(records[i].seq) <= low_limit:
            low_count[records[i].id] = len(records[i].seq)
            ##records[i].id = records[i].id + "_filtered"
            low_reads.append(records[i])
        elif len(records[i].seq) >= upper_limit:
            high_count[records[i].id] = len(records[i].seq)
            ##records[i].id = records[i].id + "_filtered"
            high_reads.append(records[i])
            if len(records[i].seq) >= highest:
                highest = len(records[i].seq)
        else:
            normal_count[records[i].id] = len(records[i].seq)
            ##records[i].id = records[i].id + "_filtered"
            normal_reads.append(records[i])
            target.write('>%s\n' % records[i].id)
            target.write('%s\n' % records[i].seq)
    
#metadata information
meta_page['A2'] = str(infile)
meta_page['B2'] = int(low_limit)
meta_page['C2'] = int(upper_limit)
meta_page['D2'] = infile + "_filtered_" + str(low_limit) + "_" + str(upper_limit)


#compile dictionary of read lengths by increment
steps.append(0)
for i in range(1,28):
    steps.append(i*100)
for i in range(0, len(steps)):
    groups.append(str(i) + "00")
records = list(SeqIO.parse(infile, "fasta"))

for j in range(0, len(steps)):
    results.setdefault(str(groups[j] + "-" + (str(j) + "99")), {'Total':0})
    for i in range(0, len(records)):
        if len(records[i].seq) >= steps[j] and len(records[i].seq) <= (steps[j]+99):
            results[groups[j] + "-" + (str(j)+ "99")]['Total'] += int(1)
results.setdefault("2800+", {'Total':0})
mid = int(highest/10)
for j in range(1,11):
    results.setdefault(str(2800 + (mid * j)) + "+", {'Total':0})
    for i in range(0, len(records)):
        if len(records[i].seq) >= 2800 + (mid * j) and len(records[i].seq) <= 2800 + (mid * (j +1)):
            results[str(2800 + (mid * j)) + "+"]['Total'] += int(1)
for i in range(0, len(records)):
    if len(records[i].seq) >2800 and len(records[i].seq) < 2800 + (highest/10):
        results['2800+']['Total'] += int(1)

#create result sheet in excel metadata workbook
meta_result = meta.create_sheet(index=1, title="Result Data")
meta_result['A1'] = "ID"
meta_result['B1'] = "Interval"
meta_result['C1'] = "Read Count"
for i, (key, value) in enumerate(results.items()):
    meta_result['A' + str(i+2)] = i
    meta_result['B' + str(i+2)] = key
    meta_result['C' + str(i+2)] = value['Total']

#Generate Chart with read length increment data
data = openpyxl.chart.Reference(meta_result, min_col=3, min_row=1, max_row=(len(steps) +2))
titles = openpyxl.chart.Reference(meta_result, min_col=2, min_row=2, max_row=(len(steps) +2))
chart = BarChart()
chart.style = 11
chart.type = 'bar'
chart.title = "Read Count by Grouping"
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)
meta_result.add_chart(chart, 'D2')

#print results to console
print("Total Read Count:" + str(len(records)))
print("Highest Read: " + str(highest))
pprint(results)
pprint("Low reads: " + str(len(low_count)))
##pprint(low_count)
pprint("Normal reads: " + str(len(normal_count)))
##pprint(normal_count)
pprint("High reads: " + str(len(high_count)))
##pprint(high_count)

#save data to ouput files. windows then linux
meta.save(infile + "_filtered_" + str(low_limit) + "_" + str(upper_limit) + "_metadata.xlsx")


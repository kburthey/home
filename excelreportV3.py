import openpyxl
import docx
import sys
import os
from openpyxl.styles import Font, PatternFill
'''
PI ID & Title: Col 4,
Activity ID: Col 5,
PI ID: Col 6,
TK ID: Col 7,
Work Type: Col 8,
Crop: Col 13,
Planned SD: Col 16,
Allocated SD: Col 17,
Planned $: Col 20,
Allocated $: Col 21,
Named Resource: Col 22,
Activity Comments: Col 23,
Portfolio: Col 24
'''

sourceFile = sys.argv[1]

#named resource report
nameSource = openpyxl.load_workbook(sourceFile)
sheets = nameSource.get_sheet_names()
source = sheets[0]
page = nameSource.get_sheet_by_name(source)
rowMax = page.max_row
staff = ['Beauchamps, Charles', 'Briois, Sebastien', 'Burthey, Kirkland', 'Cohn, Jonathan', 'Conte, Matthieu', 'El Mrani, Aicha',
         'Ganko, Eric', 'Graham, Steve', 'Green, Julie', 'Kavanaugh, Laura', 'Leysens, Maurice',  'Moughamer, Todd', 'Muller, Cedric',
         'NGOU, KISSI', 'Parvathaneni, Usha', 'Rose, Mark', 'Sharma, Pranav', 'Tiberi, Guillaume', 'Whalen, Brian']
ext = ['NCGR', 'Nebion']

#new blank report
kirkwb = openpyxl.Workbook()
newsheet = kirkwb.get_sheet_names()
kirksource = newsheet[0]
kirkpage = kirkwb.get_sheet_by_name(kirksource)

#Headers
staffTotal = []
count = 2
kirkpage['A1'] = str(page.cell(row=1, column=5).value)
kirkpage['A1'].font = Font(bold=True)
kirkpage['B1'] = str(page.cell(row=1, column=6).value)
kirkpage['B1'].font = Font(bold=True)
kirkpage['C1'] = str(page.cell(row=1, column=7).value)
kirkpage['C1'].font = Font(bold=True)
kirkpage['D1'] = str(page.cell(row=1, column=8).value)
kirkpage['D1'].font = Font(bold=True)
kirkpage['E1'] = str(page.cell(row=1, column=17).value)
kirkpage['E1'].font = Font(bold=True)
kirkpage['F1'] = str(page.cell(row=1, column=21).value)
kirkpage['F1'].font = Font(bold=True)
kirkpage['G1'] = str(page.cell(row=1, column=23).value)
kirkpage['G1'].font = Font(bold=True)

#Employee loop
for j in range(0, len(staff)):
    kirkrowMax = kirkpage.max_row
    count = kirkrowMax + 1
    employeeTotal = []
    for i in range(1,rowMax):
        if str(page.cell(row=i, column=22).value) == staff[j]:
            count += 1
            if str(page.cell(row=i, column=17).value) != '-':
                employeeTotal.append(page.cell(row=i, column=17).value)
                staffTotal.append(page.cell(row=i, column=17).value)
            if str(page.cell(row=i, column=21).value) == '-':
                kirkpage['F'+ str(count)] = str('0')
            else:
                kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
            #count += 1
            kirkpage['A' + str(kirkrowMax + 1)] = staff[j]
            kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['G' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
            kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
            kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
            kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
            kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
            #kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
            kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
    kirkpage['E' + str(kirkrowMax + 1)] = sum(employeeTotal)
staffgrandTotal = sum(staffTotal)
kirkrowMax = kirkpage.max_row
kirkpage['A' + str(kirkrowMax + 1)] = 'FTE Total (Without Not Allocated)'
kirkpage['E' + str(kirkrowMax + 1)] = str(staffgrandTotal)
kirkpage['F' + str(kirkrowMax + 1)] = str('0')
kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['G' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')

#external provider loop
providerTotal = []
for k in range(0, len(ext)):
    kirkrowMax = kirkpage.max_row
    count = kirkrowMax + 1
    extTotal = []
    for i in range(1,rowMax):
        if str(page.cell(row=i, column=22).value) == ext[k]:
            if str(page.cell(row=i, column=21).value) != '-':
                extTotal.append(page.cell(row=i, column=21).value)
                providerTotal.append(page.cell(row=i, column=21).value)
            count += 1
            kirkpage['A' + str(kirkrowMax + 1)] = ext[k]
            kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['G' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
            kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
            kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
            kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
            kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
            kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
            kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
            kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
    kirkpage['F' + str(kirkrowMax + 1)] = sum(extTotal)
providergrandTotal = sum(providerTotal)
kirkrowMax = kirkpage.max_row
kirkpage['A' + str(kirkrowMax + 1)] = 'External Provider Total'
kirkpage['E' + str(kirkrowMax + 1)] = str('0')
kirkpage['F' + str(kirkrowMax + 1)] = str(providergrandTotal)
kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['G' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')

#Not Allocated
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
employeeTotal = []
for i in range(2,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Not Allocated':
        count += 1
        if str(page.cell(row=i, column=16).value) != '-':
            employeeTotal.append(page.cell(row=i, column=16).value)
            staffTotal.append(page.cell(row=i, column=16).value)
        if str(page.cell(row=i, column=20).value) == '-':
            kirkpage['F'+ str(count)] = str('0')
        else:
            kirkpage['F'+ str(count)] = str(page.cell(row=i, column=20).value)
        #count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Not Allocated'
        kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
        kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
        kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
        kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
        kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
        kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
        kirkpage['G' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=16).value)
        #kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
kirkpage['E' + str(kirkrowMax + 1)] = sum(employeeTotal)
staffgrandTotal = sum(staffTotal)
kirkrowMax = kirkpage.max_row
kirkpage['A' + str(kirkrowMax + 1)] = 'Grand Total (With Not Allocated)'
kirkpage['E' + str(kirkrowMax + 1)] = str(staffgrandTotal)
kirkpage['F' + str(kirkrowMax + 1)] = str('0')
kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['G' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')

kirkpage.freeze_panes = 'A2'
kirkpage.title = "All Bioinformatics 2017"

'''
PI ID & Title: Col 4,
Activity ID: Col 5,
PI ID: Col 6,
TK ID: Col 7,
Work Type: Col 8,
Crop: Col 13,
Planned SD: Col 16,
Allocated SD: Col 17,
Planned $: Col 20,
Allocated $: Col 21,
Named Resource: Col 22,
Activity Comments: Col 23,
Portfolio: Col 24
'''


kirkwb.create_sheet(index=1, title="2017 Project Demand")
PInewsheet = kirkwb.get_sheet_names()
kirksource = PInewsheet[1]
kirkpage = kirkwb.get_sheet_by_name(kirksource)

#Headers
PI = []
PITotal = []
count = 2
kirkpage['A1'] = str(page.cell(row=1, column=5).value)
kirkpage['A1'].font = Font(bold=True)
kirkpage['B1'] = str(page.cell(row=1, column=8).value)
kirkpage['B1'].font = Font(bold=True)
kirkpage['C1'] = str(page.cell(row=1, column=22).value)
kirkpage['C1'].font = Font(bold=True)
kirkpage['D1'] = str(page.cell(row=1, column=17).value)
kirkpage['D1'].font = Font(bold=True)
kirkpage['E1'] = str(page.cell(row=1, column=21).value)
kirkpage['E1'].font = Font(bold=True)
kirkpage['F1'] = str(page.cell(row=1, column=23).value)
kirkpage['F1'].font = Font(bold=True)

sheetMax = page.max_row
for i in range(2, sheetMax ):
	if str(page.cell(row=i, column=4).value) not in PI:
                PI.append(page.cell(row=i, column=4).value)

#PI loop
AllPITotal = []
AllPIMoneyTotal = []
for j in range(0, len(PI)):
    kirkrowMax = kirkpage.max_row
    count = kirkrowMax + 1
    PITotal = []
    PIMoneyTotal = []
    for i in range(1,rowMax):
        if str(page.cell(row=i, column=4).value) == PI[j]:
            if str(page.cell(row=i, column=22).value) != 'Not Allocated':
                count += 1
                if str(page.cell(row=i, column=17).value) != '-':
                    PITotal.append(page.cell(row=i, column=17).value)
                    AllPITotal.append(page.cell(row=i, column=17).value)
                if str(page.cell(row=i, column=21).value) != '-':
                    PIMoneyTotal.append(page.cell(row=i, column=21).value)
                    AllPIMoneyTotal.append(page.cell(row=i, column=21).value)
                if str(page.cell(row=i, column=17).value) == '-':
                    kirkpage['D'+ str(count)] = str('0')
                else:
                    kirkpage['D'+ str(count)] = str(page.cell(row=i, column=17).value)
                if str(page.cell(row=i, column=21).value) == '-':
                    kirkpage['E'+ str(count)] = str('0')
                else:
                    kirkpage['E'+ str(count)] = str(page.cell(row=i, column=21).value)
                kirkpage['A' + str(kirkrowMax + 1)] = PI[j]
                kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
                kirkpage['B'+ str(count)] = str(page.cell(row=i, column=8).value)
                kirkpage['C'+ str(count)] = str(page.cell(row=i, column=22).value)
                kirkpage['F'+ str(count)] = str(page.cell(row=i, column=23).value)
                kirkpage['D' + str(kirkrowMax + 1)] = sum(PITotal)
                kirkpage['E' + str(kirkrowMax + 1)] = sum(PIMoneyTotal)
PIgrandTotal = sum(AllPITotal)
PIMoneygrandTotal = sum(AllPIMoneyTotal)
kirkrowMax = kirkpage.max_row
kirkpage['A' + str(kirkrowMax + 1)] = 'PI Total Without Not Allocated'
kirkpage['D' + str(kirkrowMax + 1)] = str(PIgrandTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(PIMoneygrandTotal)
kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')
kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ffcc00', fill_type='solid')

kirkrowMax = kirkpage.max_row
kirkpage['A' + str(kirkrowMax + 1)] = 'Not Allocated'
kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='d9d9d9', fill_type='solid')
kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='d9d9d9', fill_type='solid')
kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='d9d9d9', fill_type='solid')
kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='d9d9d9', fill_type='solid')
kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='d9d9d9', fill_type='solid')
kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='d9d9d9', fill_type='solid')


#PI Loop Not Allocated
PINAMoneyTotal = []
for j in range(0, len(PI)):
    kirkrowMax = kirkpage.max_row
    count = kirkrowMax + 1
    PITotal = []
    for i in range(1,rowMax):
        if str(page.cell(row=i, column=4).value) == PI[j]:
            if str(page.cell(row=i, column=22).value) == 'Not Allocated':
                count += 1
                if str(page.cell(row=i, column=16).value) != '-':
                    PITotal.append(page.cell(row=i, column=16).value)
                    AllPITotal.append(page.cell(row=i, column=16).value)
                if str(page.cell(row=i, column=20).value) != '-':
                    PINAMoneyTotal.append(page.cell(row=i, column=20).value)
                    AllPIMoneyTotal.append(page.cell(row=i, column=20).value)
                if str(page.cell(row=i, column=16).value) == '-':
                    kirkpage['D'+ str(count)] = str('0')
                else:
                    kirkpage['D'+ str(count)] = str(page.cell(row=i, column=16).value)
                if str(page.cell(row=i, column=20).value) == '-':
                    kirkpage['E'+ str(count)] = str('0')
                else:
                    kirkpage['E'+ str(count)] = str(page.cell(row=i, column=20).value)
                #count += 1
                kirkpage['A' + str(kirkrowMax + 1)] = PI[j]
                kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='ff9999', fill_type='solid')
                kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
                kirkpage['B'+ str(count)] = str(page.cell(row=i, column=8).value)
                kirkpage['C'+ str(count)] = str(page.cell(row=i, column=22).value)
                kirkpage['E'+ str(count)] = str(page.cell(row=i, column=20).value)
                kirkpage['F'+ str(count)] = str(page.cell(row=i, column=23).value)
                kirkpage['D' + str(kirkrowMax + 1)] = sum(PITotal)
                kirkpage['E' + str(kirkrowMax + 1)] = sum(PINAMoneyTotal)
PIgrandTotal = sum(AllPITotal)
PINAMoneygrandTotal = sum(PINAMoneyTotal)
kirkrowMax = kirkpage.max_row
kirkpage['A' + str(kirkrowMax + 1)] = 'PI Total With Not Allocated'
kirkpage['D' + str(kirkrowMax + 1)] = str(PIgrandTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(PINAMoneygrandTotal)
kirkpage['A' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['B' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['C' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['D' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['E' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')
kirkpage['F' + str(kirkrowMax + 1)].fill = PatternFill(start_color='33cccc', fill_type='solid')

kirkpage.freeze_panes = 'A2'
kirkwb.save('v2excelreport.xlsx')  

#! python3
'''
Use: fasta file parser
Author: Kirkland Burthey
Date: April 5 2018

# is for comments/notes
## is for disabled functions

'''
from argparse import ArgumentParser
import sys, openpyxl, os
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import(
    Reference,
    Series,
    BarChart3D,
    BarChart
)
from Bio import SeqIO
from pprint import pprint

#Declare argparse variables
parser = ArgumentParser(description='''
    For each ID in fasta input file, filter read count by
    lower and upper limits and output results to metadata
    and filtered_fasta files'''
                                 )
parser.add_argument("-i", "--infile", dest="infile", required=True,
                    help="fasta file to read in data")
parser.add_argument("-o", "--out", dest="out", required=True,
                    help="outfile directory location for metadata and filtered reads")
parser.add_argument("-l", "--low", dest="lower_limit", default=1000, type=int,
                    help="Enter a lower limit read count filter. Default: 1000")
parser.add_argument("-u", "--upper", dest="upper_limit", default=2700, type=int,
                    help="Enter an upper limit read count filter. Default: 2700")
args = parser.parse_args()
print(args)

#New blank metadata excel workbook
meta = openpyxl.Workbook()
meta_page = meta.create_sheet("Metadata",0)

#Headers for for sheet in workbook
meta_page['A1'] = "infile"
meta_page['B1'] = "max"
meta_page['C1'] = "min"
meta_page['D1'] = "outfile"

#Declare global variables
infile = os.path.basename(args.infile)
out = args.out
if os.path.exists(out):
    True
else:
    os.mkdir(out)
low_count = {}
low_reads = []
normal_count = {}
normal_reads = []
high_count = {}
high_reads = []
results = {}
steps = []
groups = []
highest = 0
try:
    low_limit = int(args.lower_limit)
    upper_limit = int(args.upper_limit)
except ValueError:
    print("lower and upper limits must be integers")
    sys.exit(0)
if low_limit > upper_limit:
    print("Lower limit must be greater than upper limit")
    sys.exit(0)
high_file = open("%s%s_filtered_reads_over_%s.fa" % (out, infile, upper_limit), "w")
low_file = open("%s%s_filtered_reads_below_%s.fa" % (out, infile, low_limit), "w")


#loop through each read and compare read length to upper and lower limits - write results
records = list(SeqIO.parse(infile, "fasta"))
with open(out + infile + "_filtered_" + str(low_limit) + "_" + str(upper_limit) + ".fa", 'w') as target:
    for i in range(0, len(records)):
        if len(records[i].seq) <= low_limit:
            low_count[records[i].id] = len(records[i].seq)
            ##records[i].id = records[i].id + "_filtered"
            low_reads.append(records[i])
            low_file.write('>%s\n%s\n' % (records[i].id, records[i].seq))
        elif len(records[i].seq) >= upper_limit:
            high_count[records[i].id] = len(records[i].seq)
            ##records[i].id = records[i].id + "_filtered"
            high_reads.append(records[i])
            high_file.write('>%s\n%s\n' % (records[i].id, records[i].seq))
            if len(records[i].seq) >= highest:
                highest = len(records[i].seq)
        else:
            normal_count[records[i].id] = len(records[i].seq)
            ##records[i].id = records[i].id + "_filtered"
            normal_reads.append(records[i])
            target.write('>%s\n' % records[i].id)
            target.write('%s\n' % records[i].seq)
low_file.close()
high_file.close()


#metadata information for first sheet in workbook
meta_page['A2'] = str(infile)
meta_page['B2'] = int(low_limit)
meta_page['C2'] = int(upper_limit)
meta_page['D2'] = "%s_filtered_%s_%s.fa" % (infile, low_limit, upper_limit)


#compile dictionary of read lengths by increment
steps.append(0)
for i in range(1,28):
    steps.append(i*100)
for i in range(0, len(steps)):
    groups.append(str(i) + "00")
records = list(SeqIO.parse(infile, "fasta"))
for j in range(0, len(steps)):
    results.setdefault(str(groups[j] + "-" + (str(j) + "99")), {'Total':0})
    for i in range(0, len(records)):
        if len(records[i].seq) >= steps[j] and len(records[i].seq) <= (steps[j]+99):
            results[groups[j] + "-" + (str(j)+ "99")]['Total'] += int(1)
results.setdefault("2800-", {'Total':0})
mid = int(highest/10)
for j in range(1,11):
    results.setdefault(str(2800 + (mid * j)) + "-", {'Total':0})
    for i in range(0, len(records)):
        if len(records[i].seq) >= 2800 + (mid * j) and len(records[i].seq) <= 2800 + (mid * (j +1)):
            results[str(2800 + (mid * j)) + "-"]['Total'] += int(1)
for i in range(0, len(records)):
    if len(records[i].seq) >2800 and len(records[i].seq) < 2800 + (highest/10):
        results['2800-']['Total'] += int(1)


#create result sheets in excel metadata workbook
meta_result = meta.create_sheet(index=1, title="Result Data")
meta_result['A1'] = "ID"
meta_result['B1'] = "Interval"
meta_result['C1'] = "Read Count"
for i, (key, value) in enumerate(results.items()):
    meta_result['A' + str(i+2)] = i
    meta_result['B' + str(i+2)] = key
    meta_result['C' + str(i+2)] = value['Total']
    
meta_result_normal = meta.create_sheet(index=2, title="Normal Reads")
meta_result_normal['A1'] = "ID"
meta_result_normal['B1'] = "Read Count"
for i, (key, value) in enumerate(normal_count.items()):
    meta_result_normal['A' + str(i+2)] = key
    meta_result_normal['B' + str(i+2)] = int(value)
    
meta_result_high = meta.create_sheet(index=3, title="High Reads")
meta_result_high['A1'] = "ID"
meta_result_high['B1'] = "Read Count"
for i, (key, value) in enumerate(high_count.items()):
    meta_result_high['A' + str(i+2)] = key
    meta_result_high['B' + str(i+2)] = int(value)
    
meta_result_low = meta.create_sheet(index=4, title="Low Reads")
meta_result_low['A1'] = "ID"
meta_result_low['B1'] = "Read Count"
for i, (key, value) in enumerate(low_count.items()):
    meta_result_low['A' + str(i+2)] = key
    meta_result_low['B' + str(i+2)] = int(value)


#Generate Chart with read length increment data 
data = openpyxl.chart.Reference(meta_result, min_col=3, min_row=1, max_row=(len(steps) +1))
titles = openpyxl.chart.Reference(meta_result, min_col=2, min_row=2, max_row=(len(steps) +1))
chart = BarChart()
chart.style = 11
chart.type = 'bar'
chart.title = "Read Count by Grouping"
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)
meta_result.add_chart(chart, 'E2')


#print results to console
print("Total Read Count:" + str(len(records)))
print("Highest Read: " + str(highest))
pprint(results)
pprint("Low reads: " + str(len(low_count)))
##pprint(low_count)
pprint("Normal reads: " + str(len(normal_count)))
##pprint(normal_count)
pprint("High reads: " + str(len(high_count)))
##pprint(high_count)


#save metadata to ouput file
meta.save("%s%s_filtered_%s_%s_metadata.xlsx" % (out, infile, low_limit, upper_limit))

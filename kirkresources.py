import openpyxl
import docx
import sys
import os
from openpyxl.styles import Font, PatternFill
from time import strftime
'''
PI ID & Title: Col 4,
Activity ID: Col 5,
PI ID: Col 6,
TK ID: Col 7,
Work Type: Col 8,
Crop: Col 13,
Planned SD: Col 16,
Allocated SD: Col 17,
Planned $: Col 20,
Allocated $: Col 21,
Named Resource: Col 22,
Activity Comments: Col 23,
Portfolio: Col 24
'''

sourceFile = sys.argv[1]

#named resource report
nameSource = openpyxl.load_workbook(sourceFile)
sheets = nameSource.get_sheet_names()
source = sheets[0]
page = nameSource.get_sheet_by_name(source)
rowMax = page.max_row

#new blank report
kirkwb = openpyxl.Workbook()
newsheet = kirkwb.get_sheet_names()
kirksource = newsheet[0]
kirkpage = kirkwb.get_sheet_by_name(kirksource)

header = Font(bold=True)
filling = PatternFill(start_color='ff9999', fill_type='solid')

#Enter name in the cmd line - Last First
name = ', '.join(sys.argv[2:])
kirkTotal = []
kirkext = []
count = 2
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == name:
        count += 1
        if str(page.cell(row=i, column=17).value) != '-':
            kirkTotal.append(page.cell(row=i, column=17).value)
        if str(page.cell(row=i, column=21).value) != '-':
            kirkext.append(page.cell(row=i, column=21).value)
        if str(page.cell(row=i, column=21).value) == '-':
            kirkpage['F'+ str(count)] = str('0')
        else:
            kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['A2'] = name
        kirkpage['A2'].fill = filling
        kirkpage['B2'].fill = filling
        kirkpage['C2'].fill = filling
        kirkpage['D2'].fill = filling
        kirkpage['E2'].fill = filling
        kirkpage['F2'].fill = filling
        kirkpage['G2'].fill = filling
        kirkpage['A1'] = str(page.cell(row=1, column=5).value)
        kirkpage['A1'].font = header
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B1'] = str(page.cell(row=1, column=6).value)
        kirkpage['B1'].font = header
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C1'] = str(page.cell(row=1, column=7).value)
        kirkpage['C1'].font = header
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D1'] = str(page.cell(row=1, column=8).value)
        kirkpage['D1'].font = header
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E1'] = str(page.cell(row=1, column=17).value)
        kirkpage['E1'].font = header
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F1'] = str(page.cell(row=1, column=21).value)
        kirkpage['F1'].font = header
        kirkpage['G1'] = str(page.cell(row=1, column=23).value)
        kirkpage['G1'].font = header
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
        
kirkgrandTotal = sum(kirkTotal)
kirkpage['E2'] = str(kirkgrandTotal)
kirkextgrandTotal = sum(kirkext)
kirkpage['F2'] = str(kirkextgrandTotal)
kirkpage.freeze_panes = 'A2'

location = os.getcwd()
kirkwb.save(location + '\\' + sys.argv[2] + 'resources (' + strftime('%m-%d-%y')+').xlsx')

from time import sleep
from datetime import datetime
from datetime import date
import openpyxl,sys
#this script reads in the Bins report and outputs the days until/passed the result delivery date

doc = sys.argv[1]
wb = openpyxl.load_workbook(doc)
sheet = wb.active
print(" ")
print ("Reading Data")
sleep (2)
print(" ")
for row in range(2, sheet.max_row +1):
	if sheet.cell(row=row, column=8).value == "active":
		z = ""
		z +=(str(sheet.cell(row=row, column=9).value))
		h=(z.split('-'))
		a = date.today()
		b = a.replace(day=int(h[1]), month=int(h[0]), year=int(h[2]))
		print (str(sheet.cell(row=row, column=1).value) + ' ' + str(b-a ))

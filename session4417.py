Python 3.6.0 (v3.6.0:41df79263a11, Dec 23 2016, 07:18:10) [MSC v.1900 32 bit (Intel)] on win32
Type "copyright", "credits" or "license()" for more information.
>>> import csv
>>> import openpyxl
>>> import docx
>>> import sys
>>> import os
>>> import re
>>> studyIdRegex = re.compile(r'\w\w\w\d{3,4}')
>>> mo = studyIdRegex.search('The study id is DAN2452')
>>> mo.group()
'DAN2452'
>>> spamfile = open('DAN2452_NCGRtoUSRE_020917.csv')
>>> spamreader = csv.reader(spamfile)
>>> spamlist = list(spamreader)
>>> spamlist
[['source', 'destination'], ['/DAN2452-chr13-minsmp5-trait-linked-region.10_20_2uniq.vcf', 'demeter:/project/DAN2452/DAN2452-chr13-minsmp5-trait-linked-region.10_20_2uniq.vcf'], ['/DAN2452-parents-CHROMOSOMES.10_20_2uniq.vcf', 'demeter:/project/DAN2452/DAN2452-parents-CHROMOSOMES.10_20_2uniq.vcf']]
>>> spamtext = ''
>>> for i in spamlist:
	spamtext += i

	
Traceback (most recent call last):
  File "<pyshell#16>", line 2, in <module>
    spamtext += i
TypeError: must be str, not list
>>> for i in spamlist:
	spamtext += str(i)

	
>>> spamtext
"['source', 'destination']['/DAN2452-chr13-minsmp5-trait-linked-region.10_20_2uniq.vcf', 'demeter:/project/DAN2452/DAN2452-chr13-minsmp5-trait-linked-region.10_20_2uniq.vcf']['/DAN2452-parents-CHROMOSOMES.10_20_2uniq.vcf', 'demeter:/project/DAN2452/DAN2452-parents-CHROMOSOMES.10_20_2uniq.vcf']"
>>> studyIdRegex = re.compile(r'[/]\w\w\w\d{3,4}')
>>> mo = studyIdRegex.search('The study id is /DAN2452')
>>> mo.groups()
()
>>> studyIdRegex = re.compile(r'\w\w\w\d{3,4}')
>>> mo = studyIdRegex.search('The study id is DAN2452')
>>> mo.groups()
()
>>> studyIdRegex = re.compile(r'[/]\w\w\w\d{3,4}')
>>> mo = studyIdRegex.search('The study id is /DAN2452')
>>> mo.group()
'/DAN2452'
>>> clarity = ''
>>> len(mo.group())
8
>>> for i in range(1,len(mo.group()):
	       
SyntaxError: invalid syntax
>>> for i in range(1, len(mo.group()):
	       
SyntaxError: invalid syntax
>>> for i in range(1, len(mo.group()):
	       
SyntaxError: invalid syntax
>>> molength = len(mo.group())
>>> molength
8
>>> for i in range(1, molength):
	clarity += mo.group()

	
>>> clarity
'/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452'
>>> for i in range(1, molength):
	clarity += str(mo.group())

	
>>> clarity
'/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452'
>>> clarity = ''
>>> for i in range(1, molength):
	clarity += str(mo.group())

	
>>> clarity
'/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452/DAN2452'
>>> for i in range(1, molength):
	clarity += str(mo.group(1:))
	
SyntaxError: invalid syntax
>>> 
>>> for i in range(1, molength):
	clarity += mo.group(1:)
	
SyntaxError: invalid syntax
>>> studyIdRegex = re.compile(r'([/])(\w\w\w\d{3,4})')
>>> mo = studyIdRegex.search('The study id is /DAN2452')
>>> mo.group()
'/DAN2452'
>>> mo.group(2)
'DAN2452'
>>> mo = studyIdRegex.search('The study id is /DAN245')
>>> mo.group()
'/DAN245'
>>> spamfile
<_io.TextIOWrapper name='DAN2452_NCGRtoUSRE_020917.csv' mode='r' encoding='cp1252'>
>>> spamreader
<_csv.reader object at 0x03A881B0>
>>> for i in spamreader:
	print i[1]
	
SyntaxError: Missing parentheses in call to 'print'
>>> for i in spamreader:
	print (i[1])

	
>>> for i in spamreader:
	print (i[0])

	
>>> spamfile = open('DAN2452_NCGRtoUSRE_020917.csv')
>>> spamreader = csv.reader(spamfile)
>>> for i in spamreader:
	print (i[1])

	
destination
demeter:/project/DAN2452/DAN2452-chr13-minsmp5-trait-linked-region.10_20_2uniq.vcf
demeter:/project/DAN2452/DAN2452-parents-CHROMOSOMES.10_20_2uniq.vcf
>>> import openpyxl
>>> wb
Traceback (most recent call last):
  File "<pyshell#71>", line 1, in <module>
    wb
NameError: name 'wb' is not defined
>>> wb = openpyxl.Workbook()
>>> sheet = wb.active
>>> for i in range(1,11):
	sheet['A' + str(i)] = i

	
>>> refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
>>> seriesObj = openpyxl.chart.Series(refObj, title="First Series")
>>> chartObj = openpyxl.chart.BarChart()
>>> chartObj.title = 'My Chart'
>>> charObj.append(seriesObj)
Traceback (most recent call last):
  File "<pyshell#81>", line 1, in <module>
    charObj.append(seriesObj)
NameError: name 'charObj' is not defined
>>> chartObj.append(seriesObj)
>>> sheet.add_chart(chartObj, 'C5')
>>> wb.save('samplechart.xlsx')
>>> names = ['Mark', 'Kirk', 'Laura', 'Julie', 'Josh', 'Matt', 'Eric', 'Steve', 'Maury', 'Todd']
>>> for i in range(1,11):
	for j in names:
		sheet['B'+str(i)] = j

		
>>> charObj.append(seriesObj)
Traceback (most recent call last):
  File "<pyshell#90>", line 1, in <module>
    charObj.append(seriesObj)
NameError: name 'charObj' is not defined
>>> chartObj.append(seriesObj)
>>> sheet.add_chart(chartObj, 'C5')
>>> wb.save('samplechart.xlsx')

Warning (from warnings module):
  File "C:\Users\t862537\Python\Python36-32\lib\zipfile.py", line 1348
    return self._open_to_write(zinfo, force_zip64=force_zip64)
UserWarning: Duplicate name: 'xl/charts/chart2.xml'
>>> wb.save('samplechart2.xlsx')
>>> sheet['B2'].value
'Todd'
>>> for i in range(1,11):
	for j in names:
		sheet['B'+str(i)] = names[j]

		
Traceback (most recent call last):
  File "<pyshell#97>", line 3, in <module>
    sheet['B'+str(i)] = names[j]
TypeError: list indices must be integers or slices, not str
>>> for i in range(1,11):
	for j in names:
		sheet['B'+str(i)] = j

		
>>> for i in range(1,11):
	for j in names:
		sheet['B'+str(i)] = j
		print(sheet['B'+str(i)].value)

		
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
>>> sheet['B2'].value
'Todd'
>>> sheet['B3'].value
'Todd'
>>> for i in range(1,11):
	for j in names:
		sheet['B'+str(i)] = j
	print(sheet['B'+str(i)].value)

	
Todd
Todd
Todd
Todd
Todd
Todd
Todd
Todd
Todd
Todd
>>> for i in range(1,11):
	for j in range(0, len(names)):
		sheet['B'+str(i)] = names[j]
	print(sheet['B'+str(i)].value)

	
Todd
Todd
Todd
Todd
Todd
Todd
Todd
Todd
Todd
Todd
>>> names
['Mark', 'Kirk', 'Laura', 'Julie', 'Josh', 'Matt', 'Eric', 'Steve', 'Maury', 'Todd']
>>> names[2]
'Laura'
>>> for i in range(1,11):
	for j in range(0, len(names)):
		sheet['B'+str(i)] = names[i]
	print(sheet['B'+str(i)].value)

	
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Traceback (most recent call last):
  File "<pyshell#112>", line 3, in <module>
    sheet['B'+str(i)] = names[i]
IndexError: list index out of range
>>> for i in range(1,11):
	for j in range(0, len(names)+1):
		sheet['B'+str(i)] = names[i]
	print(sheet['B'+str(i)].value)

	
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
Traceback (most recent call last):
  File "<pyshell#114>", line 3, in <module>
    sheet['B'+str(i)] = names[i]
IndexError: list index out of range
>>> for i in range(1,11):
	for j in range(0, len(names)):
		sheet['B'+str(i)] = names[i-1]
	print(sheet['B'+str(i)].value)

	
Mark
Kirk
Laura
Julie
Josh
Matt
Eric
Steve
Maury
Todd
>>> sheet['B2'].value
'Kirk'
>>> 
=============================== RESTART: Shell ===============================
>>> import openpyxl
>>> wb
Traceback (most recent call last):
  File "<pyshell#119>", line 1, in <module>
    wb
NameError: name 'wb' is not defined
>>> wb = openpyxl.Workbook()
>>> sheet = wb.active
>>> names = ['Mark', 'Kirk', 'Laura', 'Julie', 'Josh', 'Matt', 'Eric', 'Steve', 'Maury', 'Todd']
>>> for i in range(1,11):
	sheet['A'+str(i)] = i
	sheet['B'+str(i)] = names[i]

	
Traceback (most recent call last):
  File "<pyshell#126>", line 3, in <module>
    sheet['B'+str(i)] = names[i]
IndexError: list index out of range
>>> for i in range(1,11):
	sheet['A'+str(i)] = i
	sheet['B'+str(i)] = names[i-1]

	
>>> sheet['A2'].value
2
>>> sheet['B2'].value
'Kirk'
>>> refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
>>> seriesObj = openpyxl.chart.Series(refObj, title="First Series")
>>> chartObj = openpyxl.chart.BarChart()
>>> chartObj.title = 'My Chart'
>>> chartObj.append(seriesObj)
>>> sheet.add_chart(chartObj, 'C5')
>>> wb.save('samplechart2.xlsx')
>>> refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=2, max_row=10)
>>> seriesObj = openpyxl.chart.Series(refObj, title="First Series")
>>> chartObj = openpyxl.chart.BarChart()
>>> chartObj.title = 'My Chart'
>>> chartObj.append(seriesObj)
>>> sheet.add_chart(chartObj, 'C5')
>>> wb.save('samplechart3.xlsx')
>>> 
=============================== RESTART: Shell ===============================
>>> import openpyxl
>>> wb = openpyxl.Workbook()
>>> sheet = wb.active
>>> names = ['Mark', 'Kirk', 'Laura', 'Julie', 'Josh', 'Matt', 'Eric', 'Steve', 'Maury', 'Todd']
>>> for i in range(1,11):
	sheet['A'+str(i)] = i
	sheet['B'+str(i)] = names[i-1]

	
>>> refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
>>> seriesObj = openpyxl.chart.Series(refObj, title="First Series")
>>> chartObj = openpyxl.chart.BarChart()
>>> chartObj.title = 'My Chart'
>>> chartObj.append(seriesObj)
>>> ref2Obj = openpyxl.chart.Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=10)
>>> series2Obj = openpyxl.chart.Series(refObj, title="second Series")
>>> chartObj.append(series2Obj)
>>> sheet.add_chart(chartObj, 'C5')
>>> wb.save('samplechart3.xlsx')
>>> 
=============================== RESTART: Shell ===============================
>>> import openpyxl
>>> 
>>> wb = openpyxl.Workbook()
>>> sheet = wb.active
>>> names = ['Mark', 'Kirk', 'Laura', 'Julie', 'Josh', 'Matt', 'Eric', 'Steve', 'Maury', 'Todd']
>>> for i in range(1,11):
	sheet['A'+str(i)] = i
	sheet['B'+str(i)] = names[i-1]

	
>>> refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=2, max_row=10)
>>> seriesObj = openpyxl.chart.Series(refObj, title="First Series")
>>> chartObj = openpyxl.chart.BarChart()
>>> chartObj.title = 'My Chart'
>>> chartObj.append(seriesObj)
>>> sheet.add_chart(chartObj, 'C5')
>>> wb.save('samplechart3.xlsx')
>>> 
=============================== RESTART: Shell ===============================
>>> import openpyxl
>>> 
=============================== RESTART: Shell ===============================
>>> from openpyxl import Workbook
>>> from openpyxl.chart import (Referece, Series, BarChart3D,)
Traceback (most recent call last):
  File "<pyshell#177>", line 1, in <module>
    from openpyxl.chart import (Referece, Series, BarChart3D,)
ImportError: cannot import name 'Referece'
>>> from openpyxl.chart import (Reference, Series, BarChart3D,)
>>> wb = Workbook()
>>> ws = wb.active
>>> rows =[(None, 2013, 2014), ('Apple
			    
SyntaxError: EOL while scanning string literal
>>> rows =[(None, 2013, 2014), ('Apples', 5, 4), ('Oranges', 6, 2),('Pears', 8, 3)]
>>> rows
[(None, 2013, 2014), ('Apples', 5, 4), ('Oranges', 6, 2), ('Pears', 8, 3)]
>>> rows[2]
('Oranges', 6, 2)
>>> for row in rows:
	ws.append(row)

	
>>> ws['A2'].value
'Apples'
>>> data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=4)
>>> titles = Reference(ws, min_col=1, min_row=2, max_row=4)
>>> chart = BarChart3D()
>>> chart.title = "3D Bar Chart"
>>> chart.add_data(data=data, titles_from_data=True)

>>> ws.add_chart(chart, 'E5')
>>> wb.save('samplechart4.xlsx')
>>> 
=============================== RESTART: Shell ===============================
>>> from openpyxl import Workbook
>>> from openpyxl.chart import (Reference, Series, BarChart3D)
>>> wb = load_workbook('named_resource_breakdown4.xlsx')
Traceback (most recent call last):
  File "<pyshell#198>", line 1, in <module>
    wb = load_workbook('named_resource_breakdown4.xlsx')
NameError: name 'load_workbook' is not defined
>>> import openpyxl
>>> wb = openpyxl.load_workbook('named_resource_breakdown4.xlsx')
>>> ws = wb.active
>>> ws.max_row
586
>>> for i in range(1, ws.max_row):
	print i
	
SyntaxError: Missing parentheses in call to 'print'
>>> for i in range(1, ws.max_row):
	print (i)

	
1
2
3
4
5
6
7
8
9
10
11
12
13
14
15
16
17
18
19
20
21
22
23
24
25
26
27
28
29
30
31
32
33
34
35
36
37
38
39
40
41
42
43
44
45
46
47
48
49
50
51
52
53
54
55
56
57
58
59
60
61
62
63
64
65
66
67
68
69
70
71
72
73
74
75
76
77
78
79
80
81
82
83
84
85
86
87
88
89
90
91
92
93
94
95
96
97
98
99
100
101
102
103
104
105
106
107
108
109
110
111
112
113
114
115
116
117
118
119
120
121
122
123
124
125
126
127
128
129
130
131
132
133
134
135
136
137
138
139
140
141
142
143
144
145
146
147
148
149
150
151
152
153
154
155
156
157
158
159
160
161
162
163
164
165
166
167
168
169
170
171
172
173
174
175
176
177
178
179
180
181
182
183
184
185
186
187
188
189
190
191
192
193
194
195
196
197
198
199
200
201
202
203
204
205
206
207
208
209
210
211
212
213
214
215
216
217
218
219
220
221
222
223
224
225
226
227
228
229
230
231
232
233
234
235
236
237
238
239
240
241
242
243
244
245
246
247
248
249
250
251
252
253
254
255
256
257
258
259
260
261
262
263
264
265
266
267
268
269
270
271
272
273
274
275
276
277
278
279
280
281
282
283
284
285
286
287
288
289
290
291
292
293
294
295
296
297
298
299
300
301
302
303
304
305
306
307
308
309
310
311
312
313
314
315
316
317
318
319
320
321
322
323
324
325
326
327
328
329
330
331
332
333
334
335
336
337
338
339
340
341
342
343
344
345
346
347
348
349
350
351
352
353
354
355
356
357
358
359
360
361
362
363
364
365
366
367
368
369
370
371
372
373
374
375
376
377
378
379
380
381
382
383
384
385
386
387
388
389
390
391
392
393
394
395
396
397
398
399
400
401
402
403
404
405
406
407
408
409
410
411
412
413
414
415
416
417
418
419
420
421
422
423
424
425
426
427
428
429
430
431
432
433
434
435
436
437
438
439
440
441
442
443
444
445
446
447
448
449
450
451
452
453
454
455
456
457
458
459
460
461
462
463
464
465
466
467
468
469
470
471
472
473
474
475
476
477
478
479
480
481
482
483
484
485
486
487
488
489
490
491
492
493
494
495
496
497
498
499
500
501
502
503
504
505
506
507
508
509
510
511
512
513
514
515
516
517
518
519
520
521
522
523
524
525
526
527
528
529
530
531
532
533
534
535
536
537
538
539
540
541
542
543
544
545
546
547
548
549
550
551
552
553
554
555
556
557
558
559
560
561
562
563
564
565
566
567
568
569
570
571
572
573
574
575
576
577
578
579
580
581
582
583
584
585
>>> for i in ws.columns[4]:
	print (i)

	
Traceback (most recent call last):
  File "<pyshell#209>", line 1, in <module>
    for i in ws.columns[4]:
TypeError: 'generator' object is not subscriptable
>>> ws
<Worksheet "named_resource_breakdown (4)">
>>> ws.columns[1]
Traceback (most recent call last):
  File "<pyshell#211>", line 1, in <module>
    ws.columns[1]
TypeError: 'generator' object is not subscriptable
>>> ws.rows[1]
Traceback (most recent call last):
  File "<pyshell#212>", line 1, in <module>
    ws.rows[1]
TypeError: 'generator' object is not subscriptable
>>> ws = wb.active
>>> ws.columns[1]
Traceback (most recent call last):
  File "<pyshell#214>", line 1, in <module>
    ws.columns[1]
TypeError: 'generator' object is not subscriptable
>>> 
=============================== RESTART: Shell ===============================
>>> import openpyxl
>>> wb = openpyxl.load_workbook('example.xlsx')
>>> sheet = wb.active
>>> sheet.columns[1]
Traceback (most recent call last):
  File "<pyshell#218>", line 1, in <module>
    sheet.columns[1]
TypeError: 'generator' object is not subscriptable
>>> sheet.column[1]
Traceback (most recent call last):
  File "<pyshell#219>", line 1, in <module>
    sheet.column[1]
AttributeError: 'Worksheet' object has no attribute 'column'
>>> 
=============================== RESTART: Shell ===============================
>>> import openpyxl
>>> wb = openpyxl.load_workbook('named_resource_breakdown4.xlsx')
>>> ws = wb.active
>>> wsMax = ws.max_row + 1
>>> wsMax
587
>>> PI = []
>>> 
>>> for i in range(2, wsMax):
	if str(ws.cell(row=i, column=4).value) not in PI:
		PI.append(ws.cell(row=i, column=4).value)

		
>>> PI
['PI0008902 : LETTUCE:  Develop native trait(s) resistance - Foliar(Bremia) disease(s)', 'PI0011594 : RICE: develop TGMS markers for 2-line hybrid breeding', 'PI0001098 : WHEAT : develop methodologies for optimizing discovery and implementation for Complex Traits', 'PI0001132 : WHEAT: identify native trait(s) for resistance - insect, virus and leaf disease(s) for global', 'PI0004955 : WHEAT: identify native trait(s) for resistance - Fusarium Head Blight for global', 'PI0004970 : HYBRID WHEAT : Core global activities', 'PI0005578 : Functional Improvement Platform - Bioinformatics', 'PI0008976 : VEGETABLES_CROSS_CROP: Develop native trait(s) for control of flowering - poly-flowering', 'PI0000914 : TOMATO: Develop native trait(s) resistance - Pepino mosaic virus', 'PI0008946 : VEGETABLES_CROSS_CROP/BRASSICA: Develop native trait(s) resistance - Foliar disease(s)', 'PI0013301 : RICE: identify native trait(s) for insect resistance - BPH/wBPH', 'PI0007178 : NT: Corn Discover Water Optimization Leads', 'PI0012216 : SWEET_CORN: Germplasm management', 'PI0012819 : GenEx: Biology Research', 'PI0014132 : ET:  Establish insect mode of action and mode of resistance characterization capabilities', 'PI0009493 : POT: Evaluate R-gene Stack Technology for Rust Tolerance in Soy', 'PI0013570 : HYBRID BARLEY : maintain centralised activities providing material or data for all adaptation zones', 'PI0001105 : EARLY DEV WINTER OILSEED RAPE: identify solution(s) for resistance - disease(s) for EAME', 'PI0001097 : EARLY DEV WINTER OILSEED RAPE: identify optimized solution(s) for improvement - yield for EAME', 'PI0005276 : RICE: develop technology platform for SNPs', 'PI0001131 : EARLY DEV SUNFLOWER: maintain collaborative project SUNRISE for GLOBAL', 'PI0005197 : EARLY DEV SUNFLOWER: develop platform for marker(s) - SNP for GLOBAL', 'PI0008350 : DURUM WHEAT: identify native trait(s) for resistance - Fusarium Head Blight for global', 'PI0007086 : EARLY DEV SUGARBEET: molecular breeding for GLOBAL', 'PI0009354 : EARLY DEV WINTER OILSEED RAPE: identify native trait(s) for productivity for EAME', 'PI0002065 : CPR Method Development and Cross Project Technology Platforms', 'PI0008951 : VEGETABLES_CROSS_CROP/BRASSICA: Develop native trait(s) resistance - Xcc/black rot', 'PI0013466 : WHEAT: identify native trait(s) for resistance - Winter Hardiness for global', 'PI0004456 : RICE: develop SNP markers for restorer gene(s)', 'PI0001195 : VEGETABLES_CROSS_CROP/BRASSICA: Develop native trait(s) resistance - insect', 'PI0008893 : VEGETABLES_CROSS_CROP:  Develop native trait(s) for ripening - fruit', 'PI0005325 : MALTING BARLEY: identify native trait(s) for improvement - Malting Quality for global', 'PI0012277 : EARLY DEV SUNFLOWER: identify/deploy native trait(s) for resistance - herbicide(s) for GLOBAL', 'PI0013300 : RICE: identify native trait(s) for disease resistance - False Smut', 'PI0007051 : LG: Cross-crop discover protein leads for chewing insects', 'PI0008890 : VEGETABLES_CROSS_CROP: Develop improvement in quantitative breeding - SNP/high density array', 'PI0008907 : VEGETABLES_CROSS_CROP/CUCURBITS: Develop native trait(s) resistance - Cucumber Mosaic virus', 'PI0008914 : CUCURBITS: Develop native trait (s) resistance - Podosphaera xanthii,Golovinomyces cichoracearum', 'PI0008954 : CAULIFLOWER: Develop native trait(s) for harvestable part - White/colour', 'PI0008911 : CUCURBITS :  Develop native trait(s) resistance gynoecy, male sterility and parthenocarpy', 'PI0008919 : TOMATO : Develop native trait(s) resistance - Tospovirus', 'PI0009665 : Corn: Early_Dev Breeding research for Early-Mid Temperate germplasm (TMP01, 02, 03E) for global', 'PI0013722 : Corn: Early_Dev Breeding hybrids for Mid-Late Temperate (TMP05, 06, 07) in EAME', 'PI0007362 : PEPPER:  Develop native trait(s) - seedlessness', 'PI0015124 : PEPPER : Develop native trait(s) resistance - Oidium spp and Leveillula spp/powdery mildew', 'PI0012218 : PEPPER: Germplasm management', 'PI0001133 : EARLY DEV SUNFLOWER: identify native trait(s) for resistance - parasite(s)/broomrape for EAME', 'PI0000818 : Early Dev SOYBEAN: breed varieties for Yield Enhancement for global', 'PI0008619 : Functional Improvement Platform - Biology Technology Platforms', 'PI0015118 : TOMATO : Develop native trait(s) resistance - Xcc/bacteria leaf', 'PI0008903 : VEGETABLES_CROSS_CROP/BRASSICA: Develop native trait(s) resistance - Soil-borne/disease(s)', 'PI0008974 : OKRA: Develop native trait(s) resistance - Yellow vein mosaic/leaf curl virus', 'PI0008284 : ET: Deliver step change haploid induction rate for DH cost savings and IP for global', 'PI0014352 : Functional Improvement Platform - Genome Editing Program', 'PI0001427 : ET: Cross-crop broaden IP position on invention disclosures', 'PI0009746 : Corn-Econ SE: identify native marker(s) for - inbred producibility/male inbred fertility for global', 'PI0005999 : BRASSICA: Discover, validate & deploy genetic information in breeding', 'PI0011118 : SQUASH: Develop native trait(s) resistance - Potyvirus', 'PI0000937 : TOMATO: Develop native trait(s) for fruit quality', 'PI0001092 : WHEAT: identify native trait(s) for improvement - Grain Quality for global', 'PI0014600 : Corn: develop platform for - Marker Prediction to /Enhance Global Germplasm Exchange for global', 'PI0008960 : VEGETABLE_CROSS_CROP/CUCURBITS: Develop native trait(s) resistance-Begomovirus ', 'PI0013298 : RICE: identify native trait(s) for disease resistance - BLB', 'PI0012219 : CUCURBITS: Germplasm management', 'PI0013299 : RICE: identify native trait(s) for disease resistance - Blast', 'PI0013715 : Corn: Early_Dev Breeding hybrids for Early Dual Purpose (TMP 01,02) in EAME', 'PI0013716 : Corn: Early_Dev Breeding hybrids for Early-Mid Temperate (TMP03) in EAME', 'PI0014756 : Corn: Convert germplasm to - Cytoplasmic Male Sterility for global', 'PI0008889 : VEGETABLES_CROSS_CROP: Manage data integration - in-silico discovery', 'PI0013306 : RICE: develop genetic understanding and tools for managing photosensitivity', 'PI0009907 : ET: Cross-crop develop epigenetic approaches to enhance yield', 'PI0005345 : WHEAT: identify native trait(s) for tolerance - heat and drought stress for global', 'PI0001033 : TOMATO : Develop native trait(s) resistance - Ralstonia solanacearum', 'PI0005986 : PEPPER: Discover, validate & deploy genetic information in breeding - Hot ', 'PI0001009 : TOMATO: Develop native trait(s) resistance - Alternaria solani/fungus', 'PI0014807 : Corn: develop GM molecular stack trait for - broad Leps (locus 2) for global', 'PI0013368 : Global Fungicide Resistance CP/SC', 'PI0012429 : R&D Support to P&S Traits Assay Development and Applications', 'PI0008917 : CUCUMBER: Develop native trait(s) resistance - Corynespora cassiicola/leaf spot', 'PI0000829 : Corn: develop genetic lines and alleles for - drought tolerance/(water optimization NT) for global', 'PI0013712 : Early research on new ideas for Pasteuria based nematicides', 'PI0015204 : ET: Gene x Germplasm x Environment', 'PI0014999 : H-2016-003 (DARWIN)', 'PI0014831 : Heterosis research in support of hybrid wheat', 'PI0012217 : TOMATO: Germplasm management', 'PI0012817 : GenEx: Product Development', 'PI0015113 : PEPPER : Develop native trait(s) resistance - Botrytis cinerea/fungus']
>>> sort(PI)
Traceback (most recent call last):
  File "<pyshell#262>", line 1, in <module>
    sort(PI)
NameError: name 'sort' is not defined
>>> PI.sort()
>>> PI
['PI0000818 : Early Dev SOYBEAN: breed varieties for Yield Enhancement for global', 'PI0000829 : Corn: develop genetic lines and alleles for - drought tolerance/(water optimization NT) for global', 'PI0000914 : TOMATO: Develop native trait(s) resistance - Pepino mosaic virus', 'PI0000937 : TOMATO: Develop native trait(s) for fruit quality', 'PI0001009 : TOMATO: Develop native trait(s) resistance - Alternaria solani/fungus', 'PI0001033 : TOMATO : Develop native trait(s) resistance - Ralstonia solanacearum', 'PI0001092 : WHEAT: identify native trait(s) for improvement - Grain Quality for global', 'PI0001097 : EARLY DEV WINTER OILSEED RAPE: identify optimized solution(s) for improvement - yield for EAME', 'PI0001098 : WHEAT : develop methodologies for optimizing discovery and implementation for Complex Traits', 'PI0001105 : EARLY DEV WINTER OILSEED RAPE: identify solution(s) for resistance - disease(s) for EAME', 'PI0001131 : EARLY DEV SUNFLOWER: maintain collaborative project SUNRISE for GLOBAL', 'PI0001132 : WHEAT: identify native trait(s) for resistance - insect, virus and leaf disease(s) for global', 'PI0001133 : EARLY DEV SUNFLOWER: identify native trait(s) for resistance - parasite(s)/broomrape for EAME', 'PI0001195 : VEGETABLES_CROSS_CROP/BRASSICA: Develop native trait(s) resistance - insect', 'PI0001427 : ET: Cross-crop broaden IP position on invention disclosures', 'PI0002065 : CPR Method Development and Cross Project Technology Platforms', 'PI0004456 : RICE: develop SNP markers for restorer gene(s)', 'PI0004955 : WHEAT: identify native trait(s) for resistance - Fusarium Head Blight for global', 'PI0004970 : HYBRID WHEAT : Core global activities', 'PI0005197 : EARLY DEV SUNFLOWER: develop platform for marker(s) - SNP for GLOBAL', 'PI0005276 : RICE: develop technology platform for SNPs', 'PI0005325 : MALTING BARLEY: identify native trait(s) for improvement - Malting Quality for global', 'PI0005345 : WHEAT: identify native trait(s) for tolerance - heat and drought stress for global', 'PI0005578 : Functional Improvement Platform - Bioinformatics', 'PI0005986 : PEPPER: Discover, validate & deploy genetic information in breeding - Hot ', 'PI0005999 : BRASSICA: Discover, validate & deploy genetic information in breeding', 'PI0007051 : LG: Cross-crop discover protein leads for chewing insects', 'PI0007086 : EARLY DEV SUGARBEET: molecular breeding for GLOBAL', 'PI0007178 : NT: Corn Discover Water Optimization Leads', 'PI0007362 : PEPPER:  Develop native trait(s) - seedlessness', 'PI0008284 : ET: Deliver step change haploid induction rate for DH cost savings and IP for global', 'PI0008350 : DURUM WHEAT: identify native trait(s) for resistance - Fusarium Head Blight for global', 'PI0008619 : Functional Improvement Platform - Biology Technology Platforms', 'PI0008889 : VEGETABLES_CROSS_CROP: Manage data integration - in-silico discovery', 'PI0008890 : VEGETABLES_CROSS_CROP: Develop improvement in quantitative breeding - SNP/high density array', 'PI0008893 : VEGETABLES_CROSS_CROP:  Develop native trait(s) for ripening - fruit', 'PI0008902 : LETTUCE:  Develop native trait(s) resistance - Foliar(Bremia) disease(s)', 'PI0008903 : VEGETABLES_CROSS_CROP/BRASSICA: Develop native trait(s) resistance - Soil-borne/disease(s)', 'PI0008907 : VEGETABLES_CROSS_CROP/CUCURBITS: Develop native trait(s) resistance - Cucumber Mosaic virus', 'PI0008911 : CUCURBITS :  Develop native trait(s) resistance gynoecy, male sterility and parthenocarpy', 'PI0008914 : CUCURBITS: Develop native trait (s) resistance - Podosphaera xanthii,Golovinomyces cichoracearum', 'PI0008917 : CUCUMBER: Develop native trait(s) resistance - Corynespora cassiicola/leaf spot', 'PI0008919 : TOMATO : Develop native trait(s) resistance - Tospovirus', 'PI0008946 : VEGETABLES_CROSS_CROP/BRASSICA: Develop native trait(s) resistance - Foliar disease(s)', 'PI0008951 : VEGETABLES_CROSS_CROP/BRASSICA: Develop native trait(s) resistance - Xcc/black rot', 'PI0008954 : CAULIFLOWER: Develop native trait(s) for harvestable part - White/colour', 'PI0008960 : VEGETABLE_CROSS_CROP/CUCURBITS: Develop native trait(s) resistance-Begomovirus ', 'PI0008974 : OKRA: Develop native trait(s) resistance - Yellow vein mosaic/leaf curl virus', 'PI0008976 : VEGETABLES_CROSS_CROP: Develop native trait(s) for control of flowering - poly-flowering', 'PI0009354 : EARLY DEV WINTER OILSEED RAPE: identify native trait(s) for productivity for EAME', 'PI0009493 : POT: Evaluate R-gene Stack Technology for Rust Tolerance in Soy', 'PI0009665 : Corn: Early_Dev Breeding research for Early-Mid Temperate germplasm (TMP01, 02, 03E) for global', 'PI0009746 : Corn-Econ SE: identify native marker(s) for - inbred producibility/male inbred fertility for global', 'PI0009907 : ET: Cross-crop develop epigenetic approaches to enhance yield', 'PI0011118 : SQUASH: Develop native trait(s) resistance - Potyvirus', 'PI0011594 : RICE: develop TGMS markers for 2-line hybrid breeding', 'PI0012216 : SWEET_CORN: Germplasm management', 'PI0012217 : TOMATO: Germplasm management', 'PI0012218 : PEPPER: Germplasm management', 'PI0012219 : CUCURBITS: Germplasm management', 'PI0012277 : EARLY DEV SUNFLOWER: identify/deploy native trait(s) for resistance - herbicide(s) for GLOBAL', 'PI0012429 : R&D Support to P&S Traits Assay Development and Applications', 'PI0012817 : GenEx: Product Development', 'PI0012819 : GenEx: Biology Research', 'PI0013298 : RICE: identify native trait(s) for disease resistance - BLB', 'PI0013299 : RICE: identify native trait(s) for disease resistance - Blast', 'PI0013300 : RICE: identify native trait(s) for disease resistance - False Smut', 'PI0013301 : RICE: identify native trait(s) for insect resistance - BPH/wBPH', 'PI0013306 : RICE: develop genetic understanding and tools for managing photosensitivity', 'PI0013368 : Global Fungicide Resistance CP/SC', 'PI0013466 : WHEAT: identify native trait(s) for resistance - Winter Hardiness for global', 'PI0013570 : HYBRID BARLEY : maintain centralised activities providing material or data for all adaptation zones', 'PI0013712 : Early research on new ideas for Pasteuria based nematicides', 'PI0013715 : Corn: Early_Dev Breeding hybrids for Early Dual Purpose (TMP 01,02) in EAME', 'PI0013716 : Corn: Early_Dev Breeding hybrids for Early-Mid Temperate (TMP03) in EAME', 'PI0013722 : Corn: Early_Dev Breeding hybrids for Mid-Late Temperate (TMP05, 06, 07) in EAME', 'PI0014132 : ET:  Establish insect mode of action and mode of resistance characterization capabilities', 'PI0014352 : Functional Improvement Platform - Genome Editing Program', 'PI0014600 : Corn: develop platform for - Marker Prediction to /Enhance Global Germplasm Exchange for global', 'PI0014756 : Corn: Convert germplasm to - Cytoplasmic Male Sterility for global', 'PI0014807 : Corn: develop GM molecular stack trait for - broad Leps (locus 2) for global', 'PI0014831 : Heterosis research in support of hybrid wheat', 'PI0014999 : H-2016-003 (DARWIN)', 'PI0015113 : PEPPER : Develop native trait(s) resistance - Botrytis cinerea/fungus', 'PI0015118 : TOMATO : Develop native trait(s) resistance - Xcc/bacteria leaf', 'PI0015124 : PEPPER : Develop native trait(s) resistance - Oidium spp and Leveillula spp/powdery mildew', 'PI0015204 : ET: Gene x Germplasm x Environment']
>>> dir
<built-in function dir>
>>> dir help
SyntaxError: invalid syntax
>>> help dir
SyntaxError: invalid syntax
>>> import PyPDF2
>>> pdfFile = open('invoice.pdf', 'rb')
>>> pdfReader = pyPDF2.PdfFileReader(pdfFile)
Traceback (most recent call last):
  File "<pyshell#270>", line 1, in <module>
    pdfReader = pyPDF2.PdfFileReader(pdfFile)
NameError: name 'pyPDF2' is not defined
>>> pdfReader = PyPDF2.PdfFileReader(pdfFile)
>>> page
Traceback (most recent call last):
  File "<pyshell#272>", line 1, in <module>
    page
NameError: name 'page' is not defined
>>> pagepdf = pdfReader.getPage(0)
>>> pdfReader.numPages
1
>>> pagepdf.extractText()
''
>>> pagepdf = pdfReader.getPage(1)
Traceback (most recent call last):
  File "<pyshell#276>", line 1, in <module>
    pagepdf = pdfReader.getPage(1)
  File "C:\Users\t862537\Python\Python36-32\lib\site-packages\PyPDF2\pdf.py", line 1177, in getPage
    return self.flattenedPages[pageNumber]
IndexError: list index out of range
>>> pagepdf = pdfReader.getPage(0)
>>> pagepdf.extractText()
''
>>> pdfReader = PyPDF2.PdfFileReader(pdfFile)
>>> pagepdf = pdfReader.getPage(0)
>>> pagepdf.extractText()
''
>>> pdfFile = open('invoice.pdf', 'rb')
>>> pdfReader = PyPDF2.PdfFileReader(pdfFile)
>>> pagepdf = pdfReader.getPage(0)
>>> pagepdf.extractText()
''
>>> pdfFile = open('meetingminutes1.pdf', 'rb')
>>> pdfReader = PyPDF2.PdfFileReader(pdfFile)
>>> pagepdf = pdfReader.getPage(0)
>>> pagepdf.extractText()
'OOFFFFIICCIIAALL  BBOOAARRDD  MMIINNUUTTEESS   Meeting of \nMarch 7\n, 2014\n        \n     The Board of Elementary and Secondary Education shall provide leadership and \ncreate policies for education that expand opportunities for children, empower \nfamilies and communities, and advance Louisiana in an increasingly \ncompetitive glob\nal market.\n BOARD \n of ELEMENTARY\n and \n SECONDARY\n EDUCATION\n  '
>>> 

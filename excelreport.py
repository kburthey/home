import openpyxl
import docx
import sys
import os
'''
PI ID & Title: Col 4,
Activity ID: Col 5,
PI ID: Col 6,
TK ID: Col 7,
Work Type: Col 8,
Crop: Col 13,
Planned SD: Col 16,
Allocated SD: Col 17,
Planned $: Col 20,
Allocated $: Col 21,
Named Resource: Col 22,
Activity Comments: Col 23,
Portfolio: Col 24
'''

sourceFile = sys.argv[1]

#named resource report
nameSource = openpyxl.load_workbook(sourceFile)
sheets = nameSource.get_sheet_names()
source = sheets[0]
page = nameSource.get_sheet_by_name(source)
rowMax = page.max_row

#new blank report
kirkwb = openpyxl.Workbook()
newsheet = kirkwb.get_sheet_names()
kirksource = newsheet[0]
kirkpage = kirkwb.get_sheet_by_name(kirksource)


#first person - Charles Beauchamps
charlesTotal = []
count = 2
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Beauchamps, Charles': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        charlesTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A2'] = 'Beauchamps, Charles'
        kirkpage['A1'] = str(page.cell(row=1, column=5).value)
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B1'] = str(page.cell(row=1, column=6).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C1'] = str(page.cell(row=1, column=7).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D1'] = str(page.cell(row=1, column=8).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E1'] = str(page.cell(row=1, column=17).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F1'] = str(page.cell(row=1, column=21).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G1'] = str(page.cell(row=1, column=23).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
charlesgrandTotal = sum(charlesTotal)
kirkpage['E2'] = str(charlesgrandTotal)

#new person - Sebastien Briois
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
sebastienTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Briois, Sebastien': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        sebastienTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Briois, Sebastien'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
sebastiengrandTotal = sum(sebastienTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(sebastiengrandTotal)

#newperson - Kirk Burthey
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
kirkTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Burthey, Kirkland': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        kirkTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Burthey, Kirkland'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
kirkgrandTotal = sum(kirkTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(kirkgrandTotal)

#new person - Josh Cohn
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
joshTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Cohn, Jonathan': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        joshTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Cohn, Jonathan'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
joshgrandTotal = sum(joshTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(joshgrandTotal)

#new person - Matthieu Conte
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
matthieuTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Conte, Matthieu': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        matthieuTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Conte, Matthieu'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
matthieugrandTotal = sum(matthieuTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(matthieugrandTotal)

#new person - Aicha El Mrani
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
aichaTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'El Mrani, Aicha': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        aichaTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'El Mrani, Aicha'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
aichagrandTotal = sum(aichaTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(aichagrandTotal)

#new person - Eric Ganko
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
ericTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Ganko, Eric': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        ericTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Ganko, Eric'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
ericgrandTotal = sum(ericTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(ericgrandTotal)

#new person - Steve Graham
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
steveTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Graham, Steve': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        steveTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Graham, Steve'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
stevegrandTotal = sum(steveTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(stevegrandTotal)

#new person - Julie Green
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
julieTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Green, Julie': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        julieTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Green, Julie'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
juliegrandTotal = sum(julieTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(juliegrandTotal)

#new person - Laura Kavanaugh
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
lauraTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Kavanaugh, Laura': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        lauraTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Kavanaugh, Laura'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
lauragrandTotal = sum(lauraTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(lauragrandTotal)

#new person - Maury Leysens
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
mauryTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Leysens, Maurice': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        mauryTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Leysens, Maurice'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
maurygrandTotal = sum(mauryTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(maurygrandTotal)

#new person - Todd Moughamer
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
toddTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Moughamer, Todd': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        toddTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Moughamer, Todd'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
toddgrandTotal = sum(toddTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(toddgrandTotal)

#new person - Cedric Muller
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
cedricTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Muller, Cedric': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        cedricTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Muller, Cedric'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
cedricgrandTotal = sum(cedricTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(cedricgrandTotal)

#new person - NCGR
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
ncgrTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'NCGR': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        ncgrTotal.append(page.cell(row=i, column=21).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'NCGR'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
ncgrgrandTotal = sum(ncgrTotal)
kirkpage['F' + str(kirkrowMax + 1)] = str(ncgrgrandTotal)

#new person - Nebion
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
nebionTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Nebion': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        nebionTotal.append(page.cell(row=i, column=21).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Nebion'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
nebiongrandTotal = sum(nebionTotal)
kirkpage['F' + str(kirkrowMax + 1)] = str(nebiongrandTotal)

#new person - Kissi Ngou
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
kissiTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'NGOU, KISSI': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        if str(page.cell(row=i, column=17).value) != '-':
            kissiTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'NGOU, KISSI'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
'''kissirowMax = kirkpage.max_row
kirkpage['E' + str(kirkrowMax + 1)] = '=SUM(' + ("E" + str(kirkrowMax +2) + ":" + "E" + str(kissirowMax)) +')'   '''
kissigrandTotal = sum(kissiTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(kissigrandTotal)

#new person - Usha Parvathaneni
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
ushaTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Parvathaneni, Usha': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        ushaTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Parvathaneni, Usha'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
ushagrandTotal = sum(ushaTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(ushagrandTotal)

#new person - Mark Rose
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
markTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Rose, Mark': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        markTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Rose, Mark'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
markgrandTotal = sum(markTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(markgrandTotal)

#new person - Pranav Sharma
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
pranavTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Sharma, Pranav': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        pranavTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Sharma, Pranav'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
pranavgrandTotal = sum(pranavTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(pranavgrandTotal)

#new person - Guillaume Tiberi
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
gilTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Tiberi, Guillaume': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        gilTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Tiberi, Guillaume'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
gilgrandTotal = sum(gilTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(gilgrandTotal)

#new person - Brian Whalen
kirkrowMax = kirkpage.max_row
count = kirkrowMax + 1
brianTotal = []
for i in range(1,rowMax):
    if str(page.cell(row=i, column=22).value) == 'Whalen, Brian': #and str(page.cell(row=i, column=8).value) == 'Seq BSA':
        brianTotal.append(page.cell(row=i, column=17).value)
        count += 1
        kirkpage['A' + str(kirkrowMax + 1)] = 'Whalen, Brian'
        kirkpage['A'+ str(count)] = str(page.cell(row=i, column=5).value)
        kirkpage['B'+ str(count)] = str(page.cell(row=i, column=6).value)
        kirkpage['C'+ str(count)] = str(page.cell(row=i, column=7).value)
        kirkpage['D'+ str(count)] = str(page.cell(row=i, column=8).value)
        kirkpage['E'+ str(count)] = str(page.cell(row=i, column=17).value)
        kirkpage['F'+ str(count)] = str(page.cell(row=i, column=21).value)
        kirkpage['G'+ str(count)] = str(page.cell(row=i, column=23).value)
briangrandTotal = sum(brianTotal)
kirkpage['E' + str(kirkrowMax + 1)] = str(briangrandTotal)

kirkrowMax = kirkpage.max_row
total = kirkrowMax + 1
alltotal = []
for i in range(2, rowMax):
    if str(page.cell(row=i, column=17).value) != '-':
        alltotal.append(page.cell(row=i, column=17).value)
totaltotal = sum(alltotal)
kirkpage['A' + str(total)] = 'Total (Without Not Allocated)'
kirkpage['E' + str(total)] = str(totaltotal)



kirkwb.save('kirkexcelreport.xlsx')        



